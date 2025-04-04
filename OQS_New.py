import sys
import os
import time
import oqs
import openpyxl
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.backends import default_backend


def read_file(filepath):
    with open(filepath, 'rb') as f:
        return f.read()


def aes_encrypt(key, plaintext):
    iv = os.urandom(16)
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    padded = padder.update(plaintext) + padder.finalize()
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    ciphertext = encryptor.update(padded) + encryptor.finalize()
    return iv, ciphertext


def aes_decrypt(key, iv, ciphertext):
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    padded = decryptor.update(ciphertext) + decryptor.finalize()
    unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
    return unpadder.update(padded) + unpadder.finalize()


def process(file_path, excel_path='results.xlsx'):
    data = read_file(file_path)

    # OQS key encapsulation
    kem_alg = "Kyber512"
    with oqs.KeyEncapsulation(kem_alg) as client:
        public_key = client.generate_keypair()
        ciphertext, shared_secret_enc = client.encap_secret(public_key)

    aes_key = shared_secret_enc[:32]

    # Timing AES encryption
    start_enc = time.time()
    iv, encrypted_data = aes_encrypt(aes_key, data)
    end_enc = time.time()

    # Timing AES decryption
    start_dec = time.time()
    decrypted_data = aes_decrypt(aes_key, iv, encrypted_data)
    end_dec = time.time()

    assert data == decrypted_data, "Decryption failed: Original and decrypted data do not match."

    # Write timing results to Excel
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(["File", "Encryption Time (s)", "Decryption Time (s)"])
    else:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active

    ws.append([os.path.basename(file_path), end_enc - start_enc, end_dec - start_dec])
    wb.save(excel_path)

    # Print results
    print("\n=== Results from Excel ===")
    for row in ws.iter_rows(values_only=True):
        print(row)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 encrypt.py <file_to_encrypt.xlsx>")
        sys.exit(1)

    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"Error: File '{input_file}' does not exist.")
        sys.exit(1)

    process(input_file)
